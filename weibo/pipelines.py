# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

import copy
import csv
import html
import json
import os

import scrapy
from scrapy.exceptions import DropItem
from scrapy.pipelines.files import FilesPipeline
from scrapy.pipelines.images import ImagesPipeline
from scrapy.utils.project import get_project_settings
from weibo.cleaning_rules import evaluate_weibo

settings = get_project_settings()


def normalize_pics(pics):
    if not pics:
        return ''
    if isinstance(pics, str):
        return pics
    return ','.join(pics)


def to_int(value):
    text = str(value or '').strip()
    if not text:
        return 0
    if text.endswith('万'):
        try:
            return int(float(text[:-1]) * 10000)
        except ValueError:
            return 0
    digits = ''.join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else 0


class MergedExcelPipeline(object):
    """Merge duplicated weibos across keywords and export one cleaned workbook."""

    columns = [
        ('keywords', '关键词'),
        ('id', 'id'),
        ('bid', 'bid'),
        ('user_id', 'user_id'),
        ('screen_name', '用户昵称'),
        ('text', '微博正文'),
        ('article_url', '头条文章url'),
        ('location', '发布位置'),
        ('at_users', '艾特用户'),
        ('topics', '话题'),
        ('reposts_count', '转发数'),
        ('comments_count', '评论数'),
        ('attitudes_count', '点赞数'),
        ('created_at', '发布时间'),
        ('source', '发布工具'),
        ('pics', '微博图片url'),
        ('video_url', '微博视频url'),
        ('retweet_id', 'retweet_id'),
        ('ip', 'ip'),
        ('user_authentication', 'user_authentication'),
        ('vip_type', '会员类型'),
        ('vip_level', '会员等级'),
        ('relevance_score', '相关性分'),
        ('clean_tags', '清洗标签'),
        ('clean_reason', '清洗原因'),
    ]

    def open_spider(self, spider):
        self.items_by_id = {}
        self.total_seen = 0
        self.total_rejected = 0
        self.output_dir = settings.get('WEIBO_OUTPUT_DIR', '结果文件')
        self.excel_name = settings.get('MERGED_EXCEL_FILENAME',
                                       '水印相机行业微博.xlsx')
        self.json_name = settings.get('CLEANED_JSON_FILENAME',
                                      '水印相机行业微博.json')
        self.html_name = settings.get('DASHBOARD_FILENAME',
                                      '水印相机行业微博看板.html')
        self.export_rejected = settings.getbool('EXPORT_REJECTED_ITEMS', False)

    def process_item(self, item, spider):
        self.total_seen += 1
        keyword = item.get('keyword', '')
        weibo = dict(item['weibo'])
        weibo['pics'] = normalize_pics(weibo.get('pics', []))
        weibo_id = weibo.get('id') or weibo.get('bid')
        evaluation = evaluate_weibo(weibo, keyword)

        record = self.items_by_id.get(weibo_id)
        if not record:
            record = {
                **weibo,
                'keywords': set(),
                'relevance_score': evaluation['score'],
                'clean_tags': set(evaluation['tags']),
                'clean_reason': evaluation['reason'],
                'is_related': evaluation['is_related'],
                'matched_product_terms': set(evaluation['matched_product_terms']),
                'matched_scenario_terms': set(evaluation['matched_scenario_terms']),
                'matched_feedback_terms': set(evaluation['matched_feedback_terms']),
                'exclusion_rules': set(evaluation['exclusion_rules']),
            }
            self.items_by_id[weibo_id] = record
        else:
            record['relevance_score'] = max(record['relevance_score'],
                                            evaluation['score'])
            record['is_related'] = record['is_related'] or evaluation['is_related']
            record['clean_tags'].update(evaluation['tags'])
            record['matched_product_terms'].update(evaluation['matched_product_terms'])
            record['matched_scenario_terms'].update(evaluation['matched_scenario_terms'])
            record['matched_feedback_terms'].update(evaluation['matched_feedback_terms'])
            record['exclusion_rules'].update(evaluation['exclusion_rules'])
            if evaluation['is_related']:
                record['clean_reason'] = evaluation['reason']
        record['keywords'].add(keyword)
        return item

    def close_spider(self, spider):
        os.makedirs(self.output_dir, exist_ok=True)
        records = self.serialized_records()
        exported_records = [
            record for record in records
            if self.export_rejected or record['is_related']
        ]
        self.total_rejected = len(records) - len(exported_records)
        self.write_excel(exported_records)
        self.write_json(exported_records)
        self.write_dashboard(exported_records, records)
        spider.logger.info(
            '行业微博输出完成：总抓取%s条，合并后%s条，导出%s条，排除%s条',
            self.total_seen, len(records), len(exported_records),
            self.total_rejected)

    def serialized_records(self):
        records = []
        for record in self.items_by_id.values():
            item = dict(record)
            for key in [
                    'keywords', 'clean_tags', 'matched_product_terms',
                    'matched_scenario_terms', 'matched_feedback_terms',
                    'exclusion_rules'
            ]:
                item[key] = sorted(item.get(key, []))
            item['keywords'] = ','.join(item['keywords'])
            item['clean_tags'] = ','.join(item['clean_tags'])
            item['matched_product_terms'] = ','.join(
                item['matched_product_terms'])
            item['matched_scenario_terms'] = ','.join(
                item['matched_scenario_terms'])
            item['matched_feedback_terms'] = ','.join(
                item['matched_feedback_terms'])
            item['exclusion_rules'] = ','.join(item['exclusion_rules'])
            records.append(item)
        return sorted(
            records,
            key=lambda row: (row.get('created_at', ''), row.get('id', '')),
            reverse=True)

    def write_excel(self, records):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = '清洗后微博'
        sheet.append([label for _, label in self.columns])
        header_fill = PatternFill('solid', fgColor='1F4E78')
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for record in records:
            sheet.append([record.get(key, '') for key, _ in self.columns])

        widths = {
            'A': 28,
            'F': 70,
            'G': 28,
            'H': 14,
            'I': 24,
            'J': 24,
            'O': 50,
            'P': 50,
            'Y': 54,
        }
        for index in range(1, len(self.columns) + 1):
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = widths.get(letter, 16)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(os.path.join(self.output_dir, self.excel_name))

    def write_json(self, records):
        path = os.path.join(self.output_dir, self.json_name)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)

    def write_dashboard(self, records, all_records):
        path = os.path.join(self.output_dir, self.html_name)
        keyword_counts = {}
        for record in records:
            for keyword in record.get('keywords', '').split(','):
                if keyword:
                    keyword_counts[keyword] = keyword_counts.get(keyword, 0) + 1
        rows = []
        for record in records:
            heat = to_int(record.get('reposts_count')) + to_int(
                record.get('comments_count')) + to_int(
                    record.get('attitudes_count'))
            rows.append(f"""
            <article class="weibo-card" data-keywords="{html.escape(record.get('keywords', ''))}">
              <div class="meta">
                <span>{html.escape(record.get('created_at', ''))}</span>
                <span>{html.escape(record.get('screen_name', ''))}</span>
                <span>热度 {heat}</span>
              </div>
              <h2>{html.escape(record.get('keywords', ''))}</h2>
              <p>{html.escape(record.get('text', ''))}</p>
              <div class="tags">
                <span>{html.escape(record.get('clean_tags', '') or '行业相关')}</span>
                <span>{html.escape(record.get('clean_reason', ''))}</span>
              </div>
            </article>
            """)
        keyword_buttons = ''.join(
            f'<button type="button" data-keyword="{html.escape(keyword)}">{html.escape(keyword)} <b>{count}</b></button>'
            for keyword, count in sorted(keyword_counts.items()))
        dashboard = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>水印相机行业微博看板</title>
  <style>
    body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; color: #202124; background: #f7f8fa; }}
    header {{ padding: 24px 32px 16px; background: #ffffff; border-bottom: 1px solid #dfe3e8; }}
    h1 {{ margin: 0 0 8px; font-size: 24px; font-weight: 700; }}
    .summary {{ display: flex; gap: 16px; flex-wrap: wrap; color: #5f6368; }}
    .summary strong {{ color: #1a73e8; }}
    main {{ max-width: 1180px; margin: 0 auto; padding: 20px 24px 40px; }}
    .filters {{ display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 16px; }}
    button {{ border: 1px solid #c7d0dc; background: #ffffff; border-radius: 6px; padding: 8px 10px; cursor: pointer; }}
    button.active {{ color: #ffffff; background: #1a73e8; border-color: #1a73e8; }}
    .weibo-card {{ background: #ffffff; border: 1px solid #dfe3e8; border-radius: 8px; padding: 16px; margin-bottom: 12px; }}
    .meta, .tags {{ display: flex; gap: 10px; flex-wrap: wrap; font-size: 13px; color: #5f6368; }}
    .weibo-card h2 {{ margin: 10px 0 8px; font-size: 16px; color: #174ea6; }}
    .weibo-card p {{ margin: 0 0 12px; line-height: 1.65; white-space: pre-wrap; }}
    .tags span {{ background: #eef3fb; border-radius: 6px; padding: 5px 8px; }}
  </style>
</head>
<body>
  <header>
    <h1>水印相机行业微博看板</h1>
    <div class="summary">
      <span>清洗后 <strong>{len(records)}</strong> 条</span>
      <span>合并前 <strong>{self.total_seen}</strong> 条</span>
      <span>排除 <strong>{len(all_records) - len(records)}</strong> 条</span>
    </div>
  </header>
  <main>
    <div class="filters">
      <button class="active" type="button" data-keyword="">全部 <b>{len(records)}</b></button>
      {keyword_buttons}
    </div>
    <section id="list">
      {''.join(rows)}
    </section>
  </main>
  <script>
    const buttons = Array.from(document.querySelectorAll('button[data-keyword]'));
    const cards = Array.from(document.querySelectorAll('.weibo-card'));
    buttons.forEach(button => button.addEventListener('click', () => {{
      buttons.forEach(item => item.classList.remove('active'));
      button.classList.add('active');
      const keyword = button.dataset.keyword;
      cards.forEach(card => {{
        card.style.display = !keyword || card.dataset.keywords.includes(keyword) ? '' : 'none';
      }});
    }}));
  </script>
</body>
</html>"""
        with open(path, 'w', encoding='utf-8') as f:
            f.write(dashboard)


class CsvPipeline(object):
    def process_item(self, item, spider):
        base_dir = '结果文件' + os.sep + item['keyword']
        if not os.path.isdir(base_dir):
            os.makedirs(base_dir)
        file_path = base_dir + os.sep + item['keyword'] + '.csv'
        if not os.path.isfile(file_path):
            is_first_write = 1
        else:
            is_first_write = 0

        if item:
            with open(file_path, 'a', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                if is_first_write:
                    header = [
                        'id', 'bid', 'user_id', '用户昵称', '微博正文', '头条文章url',
                        '发布位置', '艾特用户', '话题', '转发数', '评论数', '点赞数', '发布时间',
                        '发布工具', '微博图片url', '微博视频url', 'retweet_id', 'ip', 'user_authentication',
                        '会员类型', '会员等级'
                    ]
                    writer.writerow(header)

                writer.writerow([
                    item['weibo'].get('id', ''),
                    item['weibo'].get('bid', ''),
                    item['weibo'].get('user_id', ''),
                    item['weibo'].get('screen_name', ''),
                    item['weibo'].get('text', ''),
                    item['weibo'].get('article_url', ''),
                    item['weibo'].get('location', ''),
                    item['weibo'].get('at_users', ''),
                    item['weibo'].get('topics', ''),
                    item['weibo'].get('reposts_count', ''),
                    item['weibo'].get('comments_count', ''),
                    item['weibo'].get('attitudes_count', ''),
                    item['weibo'].get('created_at', ''),
                    item['weibo'].get('source', ''),
                    normalize_pics(item['weibo'].get('pics', [])),
                    item['weibo'].get('video_url', ''),
                    item['weibo'].get('retweet_id', ''),
                    item['weibo'].get('ip', ''),
                    item['weibo'].get('user_authentication', ''),
                    item['weibo'].get('vip_type', ''),
                    item['weibo'].get('vip_level', 0)
                ])
        return item

class SQLitePipeline(object):
    def open_spider(self, spider):
        try:
            import sqlite3
            # 在结果文件目录下创建SQLite数据库
            base_dir = '结果文件'
            if not os.path.isdir(base_dir):
                os.makedirs(base_dir)
            db_name = settings.get('SQLITE_DATABASE', 'weibo.db')
            self.conn = sqlite3.connect(os.path.join(base_dir, db_name))
            self.cursor = self.conn.cursor()
            # 创建表
            sql = """
            CREATE TABLE IF NOT EXISTS weibo (
                id varchar(20) NOT NULL PRIMARY KEY,
                bid varchar(12) NOT NULL,
                user_id varchar(20),
                screen_name varchar(30),
                text varchar(2000),
                article_url varchar(100),
                topics varchar(200),
                at_users varchar(1000),
                pics varchar(3000),
                video_url varchar(1000),
                location varchar(100),
                created_at DATETIME,
                source varchar(30),
                attitudes_count INTEGER,
                comments_count INTEGER,
                reposts_count INTEGER,
                retweet_id varchar(20),
                ip varchar(100),
                user_authentication varchar(100),
                vip_type varchar(50),
                vip_level INTEGER
            )"""
            self.cursor.execute(sql)
            self.conn.commit()
        except Exception as e:
            spider.logger.error("SQLite数据库创建失败: %s", e)
            spider.sqlite3_error = True


    def process_item(self, item, spider):
        data = dict(item['weibo'])
        data['pics'] = normalize_pics(data.get('pics', []))
        keys = ', '.join(data.keys())
        placeholders = ', '.join(['?'] * len(data))
        sql = f"""INSERT OR REPLACE INTO weibo ({keys}) 
                 VALUES ({placeholders})"""
        try:
            self.cursor.execute(sql, tuple(data.values()))
            self.conn.commit()
        except Exception as e:
            spider.logger.error("SQLite保存出错: %s", e)
            spider.sqlite3_error = True
            self.conn.rollback()
        return item

    def close_spider(self, spider):
        if hasattr(self, 'conn'):
            self.conn.close()

class MyImagesPipeline(ImagesPipeline):
    def get_media_requests(self, item, info):
        if len(item['weibo']['pics']) == 1:
            yield scrapy.Request(item['weibo']['pics'][0],
                                 meta={
                                     'item': item,
                                     'sign': ''
                                 })
        else:
            sign = 0
            for image_url in item['weibo']['pics']:
                yield scrapy.Request(image_url,
                                     meta={
                                         'item': item,
                                         'sign': '-' + str(sign)
                                     })
                sign += 1

    def file_path(self, request, response=None, info=None):
        image_url = request.url
        item = request.meta['item']
        sign = request.meta['sign']
        base_dir = '结果文件' + os.sep + item['keyword'] + os.sep + 'images'
        if not os.path.isdir(base_dir):
            os.makedirs(base_dir)
        image_suffix = image_url[image_url.rfind('.'):]
        file_path = base_dir + os.sep + item['weibo'][
            'id'] + sign + image_suffix
        return file_path


class MyVideoPipeline(FilesPipeline):
    def get_media_requests(self, item, info):
        if item['weibo']['video_url']:
            yield scrapy.Request(item['weibo']['video_url'],
                                 meta={'item': item})

    def file_path(self, request, response=None, info=None):
        item = request.meta['item']
        base_dir = '结果文件' + os.sep + item['keyword'] + os.sep + 'videos'
        if not os.path.isdir(base_dir):
            os.makedirs(base_dir)
        file_path = base_dir + os.sep + item['weibo']['id'] + '.mp4'
        return file_path


class MongoPipeline(object):
    def open_spider(self, spider):
        try:
            from pymongo import MongoClient
            self.client = MongoClient(settings.get('MONGO_URI'))
            self.db = self.client['weibo']
            self.collection = self.db['weibo']
        except ModuleNotFoundError:
            spider.pymongo_error = True

    def process_item(self, item, spider):
        try:
            import pymongo

            new_item = copy.deepcopy(item)
            if not self.collection.find_one({'id': new_item['weibo']['id']}):
                self.collection.insert_one(dict(new_item['weibo']))
            else:
                self.collection.update_one({'id': new_item['weibo']['id']},
                                           {'$set': dict(new_item['weibo'])})
        except pymongo.errors.ServerSelectionTimeoutError:
            spider.mongo_error = True

    def close_spider(self, spider):
        try:
            self.client.close()
        except AttributeError:
            pass


class MysqlPipeline(object):
    def create_database(self, mysql_config):
        """创建MySQL数据库"""
        import pymysql
        sql = """CREATE DATABASE IF NOT EXISTS %s DEFAULT
            CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci""" % settings.get(
            'MYSQL_DATABASE', 'weibo')
        db = pymysql.connect(**mysql_config)
        cursor = db.cursor()
        cursor.execute(sql)
        db.close()

    def create_table(self):
        """创建MySQL表"""
        sql = """
                CREATE TABLE IF NOT EXISTS weibo (
                id varchar(20) NOT NULL,
                bid varchar(12) NOT NULL,
                user_id varchar(20),
                screen_name varchar(30),
                text varchar(2000),
                article_url varchar(100),
                topics varchar(200),
                at_users varchar(1000),
                pics varchar(3000),
                video_url varchar(1000),
                location varchar(100),
                created_at DATETIME,
                source varchar(30),
                attitudes_count INT,
                comments_count INT,
                reposts_count INT,
                retweet_id varchar(20),
                PRIMARY KEY (id),
                ip varchar(100),
                user_authentication varchar(100),
                vip_type varchar(50),
                vip_level INT
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4"""
        self.cursor.execute(sql)

    def open_spider(self, spider):
        try:
            import pymysql
            mysql_config = {
                'host': settings.get('MYSQL_HOST', 'localhost'),
                'port': settings.get('MYSQL_PORT', 3306),
                'user': settings.get('MYSQL_USER', 'root'),
                'password': settings.get('MYSQL_PASSWORD', '123456'),
                'charset': 'utf8mb4'
            }
            self.create_database(mysql_config)
            mysql_config['db'] = settings.get('MYSQL_DATABASE', 'weibo')
            self.db = pymysql.connect(**mysql_config)
            self.cursor = self.db.cursor()
            self.create_table()
        except ImportError:
            spider.pymysql_error = True
        except pymysql.OperationalError:
            spider.mysql_error = True

    def process_item(self, item, spider):
        data = dict(item['weibo'])
        data['pics'] = normalize_pics(data.get('pics', []))
        keys = ', '.join(data.keys())
        values = ', '.join(['%s'] * len(data))
        sql = """INSERT INTO {table}({keys}) VALUES ({values}) ON
                     DUPLICATE KEY UPDATE""".format(table='weibo',
                                                    keys=keys,
                                                    values=values)
        update = ','.join([" {key} = {key}".format(key=key) for key in data])
        sql += update
        try:
            self.cursor.execute(sql, tuple(data.values()))
            self.db.commit()
        except Exception:
            self.db.rollback()
        return item

    def close_spider(self, spider):
        try:
            self.db.close()
        except Exception:
            pass


class DuplicatesPipeline(object):
    def __init__(self):
        self.ids_seen = set()

    def process_item(self, item, spider):
        if item['weibo']['id'] in self.ids_seen:
            raise DropItem("过滤重复微博: %s" % item)
        else:
            self.ids_seen.add(item['weibo']['id'])
            return item
